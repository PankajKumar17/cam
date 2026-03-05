"""
Yakṣarāja — Financial Statement Parser
=============================================
Parses multi-tab Excel files (Screener.in format) and CSV files
into the company_data dict expected by the pipeline.

Screener.in Excel structure:
  - "Data Sheet" tab has ALL raw numeric data (formulas in other sheets reference it)
  - P&L rows 15-31, BS rows 55-72, CF rows 80-85
  - Each column = one fiscal year (datetime objects in the Report Date row)

Author: Team Yakṣarāja
"""

import os
import re
from datetime import datetime

import pandas as pd
import numpy as np
import openpyxl
from loguru import logger


def _clean_numeric(val):
    """Convert a cell value to float, handling commas, %, blanks, dashes."""
    if val is None or val == "" or val == "-":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("₹", "").replace("%", "")
    if s in ("", "-", "--", "N/A", "NA", "nan"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _safe_div(a, b, default=None):
    """Safe division that handles None and zero."""
    if a is None or b is None or b == 0:
        return default
    return a / b


def parse_screener_excel(filepath: str, company_name: str = None) -> dict:
    """
    Parse a Screener.in multi-tab Excel export into the company_data dict.

    Strategy:
      1. Try reading from 'Data Sheet' tab (has raw values)
      2. If no Data Sheet, try reading from formula sheets with pandas
      3. For CSV files, parse as flat table

    Args:
        filepath: Path to Excel or CSV file
        company_name: Company name (auto-detected from file if not provided)

    Returns:
        dict matching pipeline's company_data format
    """
    logger.info(f"Parsing financial file: {filepath}")

    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".csv":
        return _parse_csv(filepath, company_name)

    # ── Try Screener.in format (Data Sheet tab) ──────────────────────────
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if "Data Sheet" in wb.sheetnames:
            logger.info("Found 'Data Sheet' tab — using Screener.in parser")
            return _parse_data_sheet(wb, company_name)
        else:
            logger.info("No 'Data Sheet' tab — trying pandas-based parser")
            return _parse_formula_sheets(filepath, company_name)
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        return _empty_company_data(company_name or "Unknown")


def _parse_data_sheet(wb, company_name: str = None) -> dict:
    """
    Parse Screener.in 'Data Sheet' tab which has all raw numeric data.

    Layout:
      Row 1:  COMPANY NAME | <name>
      Row 16: Report Date  | date1 | date2 | ... (P&L dates)
      Row 17: Sales        | val1  | val2  | ...
      ...
      Row 55: BALANCE SHEET
      Row 56: Report Date  | date1 | date2 | ...
      Row 57: Equity Share Capital | ...
      ...
      Row 80: CASH FLOW:
      Row 81: Report Date  | date1 | date2 | ...
      Row 82: Cash from Operating Activity | ...
    """
    ws = wb["Data Sheet"]

    # ── Detect company name ──────────────────────────────────────────────
    if company_name is None:
        cell_val = ws.cell(1, 2).value
        company_name = str(cell_val).strip() if cell_val else "Unknown Company"

    logger.info(f"Company: {company_name}")

    # ── Build row lookup (label → row_index) ─────────────────────────────
    row_map = {}
    for row_idx in range(1, ws.max_row + 1):
        label = ws.cell(row_idx, 1).value
        if label is not None:
            row_map[str(label).strip().lower()] = row_idx

    # ── Find latest fiscal year column ───────────────────────────────────
    # Report Date row (row 16 for P&L) has datetime objects
    report_date_row = row_map.get("report date", 16)

    # Find the rightmost column with a date value
    latest_col = None
    latest_year = 0
    all_cols = []  # (col_idx, year) sorted desc

    for col_idx in range(2, ws.max_column + 1):
        val = ws.cell(report_date_row, col_idx).value
        if val is not None:
            if isinstance(val, datetime):
                yr = val.year
            elif isinstance(val, (int, float)):
                yr = int(val)
            else:
                m = re.search(r"20\d{2}", str(val))
                yr = int(m.group()) if m else 0
            if yr > 0:
                all_cols.append((col_idx, yr))
                if yr > latest_year:
                    latest_year = yr
                    latest_col = col_idx

    all_cols.sort(key=lambda x: x[1], reverse=True)
    fiscal_year = latest_year

    if latest_col is None:
        logger.warning("Could not detect year columns in Data Sheet")
        return _empty_company_data(company_name)

    logger.info(f"Latest year column: col {latest_col} → FY{fiscal_year}")

    # ── Helper to get value from a row ───────────────────────────────────
    def _get(label_variants, col=None):
        """Get numeric value for a row label at a given column."""
        if col is None:
            col = latest_col
        for label in label_variants:
            key = label.strip().lower()
            if key in row_map:
                val = ws.cell(row_map[key], col).value
                return _clean_numeric(val)
        return None

    def _get_prev(label_variants, years_back=1):
        """Get value from a previous year column."""
        if len(all_cols) > years_back:
            prev_col = all_cols[years_back][0]
            return _get(label_variants, prev_col)
        return None

    # ── P&L data ─────────────────────────────────────────────────────────
    revenue = _get(["Sales", "Revenue", "Net Sales", "Income from operations"])
    raw_material = _get(["Raw Material Cost"])
    employee_cost = _get(["Employee Cost"])
    other_expenses = _get(["Other Expenses"])
    selling_admin = _get(["Selling and admin"])
    other_income = _get(["Other Income"])
    depreciation = _get(["Depreciation"])
    interest = _get(["Interest", "Interest Expense", "Finance Cost"])
    pbt = _get(["Profit before tax", "PBT"])
    tax = _get(["Tax", "Tax Expense"])
    pat = _get(["Net profit", "PAT", "Net Profit"])

    # EBITDA = PBT + Interest + Depreciation - Other Income
    # This is more reliable than summing cost items
    if pbt is not None:
        ebitda = (pbt or 0) + (interest or 0) + (depreciation or 0) - (other_income or 0)
    elif revenue:
        # Fallback: Revenue - sum of operating cost items
        cost_items = [raw_material, _get(["Change in Inventory"]),
                      _get(["Power and Fuel"]), _get(["Other Mfr. Exp"]),
                      employee_cost, selling_admin, other_expenses]
        expenses = sum(v for v in cost_items if v is not None)
        ebitda = (revenue or 0) - expenses
    else:
        ebitda = 0

    # Previous year revenue for growth
    prev_revenue = _get_prev(["Sales", "Revenue"])

    # ── Balance Sheet data ───────────────────────────────────────────────
    # BS uses its own Report Date row (row 56)
    bs_date_row = row_map.get("balance sheet", None)
    if bs_date_row:
        # Find the BS report date row (next row after BALANCE SHEET header)
        for r in range(bs_date_row + 1, bs_date_row + 3):
            label = ws.cell(r, 1).value
            if label and "report date" in str(label).lower():
                break

    equity_capital = _get(["Equity Share Capital"])
    reserves = _get(["Reserves"])
    total_equity = ((equity_capital or 0) + (reserves or 0)) if (equity_capital is not None or reserves is not None) else None

    borrowings = _get(["Borrowings", "Total Borrowings"])
    other_liab = _get(["Other Liabilities"])

    total_assets_row1 = None
    # Find the first "Total" after Balance Sheet section
    for key, row_idx in row_map.items():
        if key == "total" and row_idx > (row_map.get("balance sheet", 50)):
            # Check if it's the liabilities total or assets total
            if total_assets_row1 is None:
                total_assets_row1 = row_idx
            else:
                break

    total_assets = _get(["Total"], col=latest_col) if total_assets_row1 else None
    # Use the value from the assets-side Total
    if total_assets_row1:
        # The Data Sheet has two "Total" rows - liabilities (row 61) and assets (row 66)
        # We need the second one (which equals the first in a proper BS)
        total_val = ws.cell(total_assets_row1, latest_col).value
        total_assets = _clean_numeric(total_val)

    net_block = _get(["Net Block", "Fixed Assets", "Total Fixed Assets"])
    receivables = _get(["Receivables", "Trade Receivables", "Debtors"])
    inventory = _get(["Inventory", "Inventories"])
    cash_bank = _get(["Cash & Bank", "Cash and Bank", "Cash and Cash Equivalents"])

    # ── Cash Flow data ───────────────────────────────────────────────────
    cfo = _get(["Cash from Operating Activity", "Cash from Operations"])
    cfi = _get(["Cash from Investing Activity"])
    cff = _get(["Cash from Financing Activity"])

    # ── Compute derived metrics ──────────────────────────────────────────
    ebitda_margin = _safe_div(ebitda, revenue)
    net_margin = _safe_div(pat, revenue)
    gross_margin = _safe_div((revenue or 0) - (raw_material or 0), revenue) if revenue else None

    debt_to_equity = _safe_div(borrowings, total_equity)
    interest_coverage = _safe_div(ebitda, interest) if interest and interest > 0 else None
    roe = _safe_div(pat, total_equity)
    roa = _safe_div(pat, total_assets)
    revenue_growth = _safe_div(((revenue or 0) - (prev_revenue or 0)), prev_revenue) if prev_revenue else None

    # DSCR = (PAT + Depreciation + Interest) / (Interest + Debt Repayment)
    debt_repayment = (borrowings or 0) * 0.15
    dscr_num = (pat or 0) + (depreciation or 0) + (interest or 0)
    dscr_den = (interest or 0) + debt_repayment
    dscr = _safe_div(dscr_num, dscr_den, default=1.5)

    # Current ratio approximation
    other_assets = _get(["Other Assets"])
    current_assets = (receivables or 0) + (inventory or 0) + (cash_bank or 0)
    current_liab = (other_liab or 0) * 0.6  # Rough estimate
    current_ratio = _safe_div(current_assets, current_liab, default=1.0)

    capex = abs(cfi) if cfi and cfi < 0 else 0
    fcf = (cfo or 0) - capex

    # Altman Z-Score
    wc = current_assets - current_liab
    ta = total_assets or 1
    ebit = ebitda - (depreciation or 0)
    altman_z = (1.2 * (wc / ta) +
                1.4 * ((total_equity or 0) * 0.5 / ta) +
                3.3 * (ebit / ta) +
                0.6 * _safe_div(total_equity, borrowings, default=1.0) +
                1.0 * _safe_div(revenue, ta, default=0.5))

    # Beneish M-Score — needs 2yr data
    beneish_m = -2.5
    prev_receivables = _get_prev(["Receivables", "Trade Receivables"])
    prev_rev = prev_revenue
    if prev_receivables and prev_rev and receivables and revenue:
        dsri = _safe_div(receivables / revenue, prev_receivables / prev_rev, default=1.0)
        beneish_m = -4.84 + 0.92 * (dsri or 1.0)  # Simplified single-variable

    # Piotroski F-Score
    f_score = 0
    if pat and pat > 0: f_score += 1
    if cfo and cfo > 0: f_score += 1
    if roa and roa > 0: f_score += 1
    if cfo and pat and cfo > pat: f_score += 1
    if debt_to_equity and debt_to_equity < 2.0: f_score += 1
    if current_ratio and current_ratio > 1.0: f_score += 1
    if ebitda_margin and ebitda_margin > 0.10: f_score += 1
    if revenue_growth and revenue_growth > 0: f_score += 1
    f_score = min(f_score, 9)

    # ── Build output dict ────────────────────────────────────────────────
    data = {
        "company_name": company_name,
        "sector": _detect_sector(company_name),
        "fiscal_year": fiscal_year,
        # P&L
        "revenue": revenue or 0,
        "ebitda": round(ebitda, 2) if ebitda else 0,
        "ebitda_margin": round(ebitda_margin, 3) if ebitda_margin else 0,
        "pat": pat or 0,
        "net_margin": round(net_margin, 3) if net_margin else 0,
        "gross_margin": round(gross_margin, 3) if gross_margin else 0,
        "depreciation": depreciation or 0,
        "interest_expense": interest or 0,
        "ebit": round(ebit, 2) if ebit else 0,
        "pbt": pbt or 0,
        "tax": tax or 0,
        "other_income": other_income or 0,
        "employee_cost": employee_cost or 0,
        # Balance sheet
        "total_assets": total_assets or 0,
        "total_equity": round(total_equity, 2) if total_equity else 0,
        "total_debt": borrowings or 0,
        "lt_borrowings": (borrowings or 0) * 0.65,
        "st_borrowings": (borrowings or 0) * 0.35,
        "fixed_assets": net_block or 0,
        "trade_receivables": receivables or 0,
        "inventories": inventory or 0,
        "cash_equivalents": cash_bank or 0,
        "total_current_assets": current_assets,
        "total_current_liab": current_liab,
        # Cash flow
        "cfo": cfo or 0,
        "cfi": cfi or 0,
        "cff": cff or 0,
        "capex": capex,
        "free_cash_flow": round(fcf, 2),
        # Ratios
        "dscr": round(dscr, 2) if dscr else 1.5,
        "interest_coverage": round(interest_coverage, 2) if interest_coverage else 2.0,
        "debt_to_equity": round(debt_to_equity, 2) if debt_to_equity else 1.5,
        "current_ratio": round(current_ratio, 2) if current_ratio else 1.0,
        "roe": round(roe, 3) if roe else 0,
        "roa": round(roa, 3) if roa else 0,
        "revenue_growth": round(revenue_growth, 3) if revenue_growth else 0,
        # Governance — leave None so CAM shows "Not Given" when real data is absent
        "promoter_holding_pct": None,
        "promoter_pledge_pct": None,
        "institutional_holding_pct": None,
        "related_party_tx_to_rev": None,
        "receivables_days": round(365 * (receivables or 0) / max(revenue or 1, 1), 0),
        # Forensic scores
        "beneish_m_score": round(beneish_m, 2),
        "beneish_dsri": 0.95,
        "beneish_tata": 0.03,
        "altman_z_score": round(altman_z, 2),
        "piotroski_f_score": f_score,
        "auditor_distress_score": 1,
        "going_concern_flag": 0,
        "qualified_opinion_flag": 0,
        "auditor_resigned_flag": 0,
        # Alt data placeholders
        "contagion_risk_score": 0.0,
        "network_npa_ratio": 0.0,
        "gst_vs_bank_divergence": 0.0,
        "satellite_activity_score": 0.0,
        "employee_cost_to_rev": _safe_div(employee_cost, revenue, default=0.15),
        "st_debt_to_lt_assets_ratio": _safe_div((borrowings or 0) * 0.35,
                                                 net_block, default=0.4),
        "cfo_to_debt": _safe_div(cfo, borrowings, default=0.15),
        "debt_growth_3yr": 0.10,
        "cfo_to_pat": _safe_div(cfo, pat, default=0.8) if pat and pat != 0 else 0.8,
        "free_cash_flow_margin": _safe_div(fcf, revenue, default=0.03),
        "label": 0,
    }

    logger.info(f"✅ Parsed {company_name}: Revenue=₹{revenue:,.0f} Cr, "
                f"EBITDA=₹{ebitda:,.0f} Cr, PAT=₹{pat:,.0f} Cr, "
                f"DSCR={data['dscr']}, D/E={data['debt_to_equity']}")

    return data


def _parse_formula_sheets(filepath: str, company_name: str = None) -> dict:
    """Fallback: try parsing formula-based sheets with pandas."""
    try:
        sheets = pd.read_excel(filepath, sheet_name=None, header=None)
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        return _empty_company_data(company_name or "Unknown")

    if company_name is None:
        first_sheet = list(sheets.values())[0]
        candidate = str(first_sheet.iloc[0, 0]).strip()
        if candidate and "nan" not in candidate.lower() and len(candidate) > 2:
            company_name = candidate
        else:
            company_name = "Unknown Company"

    logger.warning(f"No Data Sheet found — parsing may be incomplete for {company_name}")
    return _empty_company_data(company_name)


def _parse_csv(filepath: str, company_name: str = None) -> dict:
    """Parse a CSV file with financial data."""
    try:
        df = pd.read_csv(filepath)
        if company_name is None:
            company_name = "Unknown Company"

        # Try to extract key metrics from CSV columns or rows
        result = _empty_company_data(company_name)

        # Check if it's a row-based format (metrics in first column)
        if df.shape[1] >= 2:
            for _, row in df.iterrows():
                label = str(row.iloc[0]).strip().lower()
                val = _clean_numeric(row.iloc[-1])  # Use last column
                if val is not None:
                    if label in ("sales", "revenue"):
                        result["revenue"] = val
                    elif label in ("net profit", "pat"):
                        result["pat"] = val
                    elif label in ("operating profit", "ebitda"):
                        result["ebitda"] = val
                    elif label in ("borrowings", "total debt"):
                        result["total_debt"] = val
                    elif label in ("depreciation",):
                        result["depreciation"] = val
                    elif label in ("interest",):
                        result["interest_expense"] = val

        return result

    except Exception as e:
        logger.error(f"Failed to read CSV: {e}")
        return _empty_company_data(company_name or "Unknown")


def _detect_sector(company_name: str) -> str:
    """Simple sector detection from company name."""
    name = company_name.lower()
    sector_map = {
        "textile": "Textiles", "fabric": "Textiles", "yarn": "Textiles",
        "energy": "Energy", "power": "Energy", "solar": "Energy", "wind": "Energy",
        "suzlon": "Energy",
        "pharma": "Pharmaceuticals", "drug": "Pharmaceuticals", "bio": "Biotechnology",
        "bank": "Banking", "finance": "Financial Services", "nbfc": "Financial Services",
        "infra": "Infrastructure", "construction": "Infrastructure",
        "steel": "Metals & Mining", "metal": "Metals & Mining", "mining": "Metals & Mining",
        "auto": "Automobile", "motor": "Automobile", "vehicle": "Automobile",
        "tech": "Technology", "software": "Technology", "it ": "Technology",
        "food": "FMCG", "consumer": "FMCG", "fmcg": "FMCG",
        "real": "Real Estate", "housing": "Real Estate",
        "oil": "Oil & Gas", "petrol": "Oil & Gas", "gas": "Oil & Gas",
        "cement": "Cement", "chemical": "Chemicals",
    }
    for keyword, sector in sector_map.items():
        if keyword in name:
            return sector
    return "Industrial"


def _empty_company_data(company_name: str) -> dict:
    """Return a minimal company_data dict with defaults."""
    return {
        "company_name": company_name,
        "sector": _detect_sector(company_name),
        "fiscal_year": 2025,
        "revenue": 0, "ebitda": 0, "ebitda_margin": 0, "pat": 0,
        "net_margin": 0, "gross_margin": 0, "depreciation": 0,
        "interest_expense": 0, "ebit": 0, "pbt": 0, "tax": 0,
        "total_assets": 0, "total_equity": 0, "total_debt": 0,
        "lt_borrowings": 0, "st_borrowings": 0,
        "fixed_assets": 0, "trade_receivables": 0, "inventories": 0,
        "cash_equivalents": 0, "total_current_assets": 0, "total_current_liab": 0,
        "cfo": 0, "cfi": 0, "cff": 0, "capex": 0, "free_cash_flow": 0,
        "dscr": 1.5, "interest_coverage": 2.0, "debt_to_equity": 1.5,
        "current_ratio": 1.0, "roe": 0, "roa": 0, "revenue_growth": 0,
        "promoter_holding_pct": None, "promoter_pledge_pct": None,
        "institutional_holding_pct": None, "related_party_tx_to_rev": None,
        "receivables_days": 55,
        "beneish_m_score": -2.5, "beneish_dsri": 0.95, "beneish_tata": 0.03,
        "altman_z_score": 2.5, "piotroski_f_score": 5,
        "auditor_distress_score": 1, "going_concern_flag": 0,
        "qualified_opinion_flag": 0, "auditor_resigned_flag": 0,
        "contagion_risk_score": 0.0, "network_npa_ratio": 0.0,
        "gst_vs_bank_divergence": 0.0, "satellite_activity_score": 0.0,
        "employee_cost_to_rev": 0.15, "st_debt_to_lt_assets_ratio": 0.4,
        "cfo_to_debt": 0.15, "debt_growth_3yr": 0.10,
        "cfo_to_pat": 0.8, "free_cash_flow_margin": 0.03,
        "label": 0,
    }
